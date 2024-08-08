# Importing useful libraries
import OleFileIO_PL # For reading excel files
import openpyxl # For excel related tasks
import pandas as pd # For dataframe related tasks
from openpyxl.styles import PatternFill, Font, Alignment # For excel formating
from openpyxl.utils import get_column_letter
import swifter # For paralalizing long taking tasks
import zipfile # For zipping file together

# For operating system related tasks
import os 
import shutil

import numpy as np # For numerical tasks
import streamlit as st # For simple web app

# A function to zip the files
def zip_files(file_paths, output_zip_file):
    
    with zipfile.ZipFile(output_zip_file, 'w') as zipf:
        for file in file_paths:
            zipf.write(file, os.path.basename(file))

# Reading excel file
def read_excel_file(path):

    with open(path,'rb') as file:
        ole = OleFileIO_PL.OleFileIO(file)
        if ole.exists('Workbook'):
            d = ole.openstream('Workbook')
            data=pd.read_excel(d,engine='xlrd')

    return data

# Function to apply formatting to a sheet
def format_sheet(sheet, conditional_formating = False,error_sheet=False,warning_sheet=False,findings=False):
    
    # Color the first row black
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_bold_font = Font(color="FFFFFF", bold=True)

    # Merging and adding custom texts
    def merge_and_style_cells(start_row, end_row, start_col, end_col, text, fill, font_size):

        # Merging the cells
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

        # Adding values into the merged cells
        merged_cell = sheet.cell(row=start_row, column=start_col)
        merged_cell.value = text
        merged_cell.fill = fill
        merged_cell.font = Font(color="000000", size=font_size)
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Calculate the width of the text to adjust the column width
        text_length = len(text) # Getting the text length

        # Itterating through all the columns
        for col in range(start_col, end_col + 1):

            # Understanding the total column letters to adjust the size
            column_letter = get_column_letter(col)
            current_width = sheet.column_dimensions[column_letter].width
            
            if current_width is None:
                current_width = 0

            # Adjust the column width to be the maximum of current width or text length
            new_width = max(current_width, text_length + 2)
            sheet.column_dimensions[column_letter].width = new_width
    
    if findings != False:
        unique_findings,must_stop_findings,below_40_findings,exactly_40_findings,above_40_findings,errors = findings
        # Merge and style specific rows
        merge_and_style_cells(2, 5, 1, 13, unique_findings, PatternFill(start_color ="FFFF99", end_color ="FFFF99", fill_type ="solid"), 30)
        merge_and_style_cells(6, 9, 1, 13, must_stop_findings, PatternFill(start_color ="FF6666", end_color ="FF6666", fill_type ="solid"), 30)
        merge_and_style_cells(10, 13, 1, 13, below_40_findings, PatternFill(start_color ="FFCCCC", end_color ="FFCCCC", fill_type ="solid"), 30)
        merge_and_style_cells(14, 17, 1, 13, exactly_40_findings, PatternFill(start_color ="FFFF99", end_color ="FFFF99", fill_type ="solid"), 30)
        merge_and_style_cells(18, 21, 1, 13, above_40_findings, PatternFill(start_color ="CCFFCC", end_color ="CCFFCC", fill_type ="solid"), 30)
        merge_and_style_cells(22, 24, 1, 13, errors, PatternFill(start_color = "FF6666", end_color ="CCFFCC", fill_type ="solid"), 30)

    if findings == False:
        for cell in sheet[1]:
            cell.fill = black_fill
            cell.font = white_bold_font # Change font color to white for readability

    # Adjust column widths to fit the contents
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Conditional formatting based on average percentage
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    if conditional_formating:

        if conditional_formating == 'registration_rate':

            # Find the column with the header "Average Percentage"
            avg_col = None
            for cell in sheet[1]:
                if cell.value == "Average Percentage":
                    avg_col = cell.column
                    break

            if avg_col is None:
                print("Column 'Average Percentage' not found.")
                return

            for row in sheet.iter_rows(min_row=2):
                avg_cell = row[avg_col - 1]
                try:
                    value = float(avg_cell.value)
                    if value < 10:
                        row_fill = red_fill
                    elif 10 <= value < 40:
                        row_fill = yellow_fill
                    elif value >= 40:
                        row_fill = green_fill
                    
                    for cell in row:
                        cell.fill = row_fill
                except (ValueError, TypeError):
                    continue

        elif conditional_formating == 'margin':

            # Find the column with the header "Margin"
            avg_col = None
            for cell in sheet[1]:
                if cell.value == "Average Margin":
                    avg_col = cell.column
                    break

            if avg_col is None:
                print("Column 'Margin' not found.")
                return

            for row in sheet.iter_rows(min_row=2):
                avg_cell = row[avg_col - 1]
                try:
                    value = float(avg_cell.value)
                    if value < 40:
                        row_fill = red_fill
                    elif 40 <= value < 70:
                        row_fill = yellow_fill
                    elif value >= 70:
                        row_fill = green_fill
                    
                    for cell in row:
                        cell.fill = row_fill
                except (ValueError, TypeError):
                    continue

    if error_sheet:

        for row in sheet.iter_rows(min_row=2):

            for cell in row:
                cell.fill = red_fill 

    if warning_sheet:

        for row in sheet.iter_rows(min_row=2):

            for cell in row:
                cell.fill = yellow_fill

    # Coloring the cell containing the total data
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Total':
                cell.fill = black_fill
                cell.font = white_bold_font

def return_revenue(revenue_data,date,name,reve_id,sub_id):

    temp_rev_data = revenue_data[
        (revenue_data['Company'] == name) & (revenue_data['Affiliate'] == reve_id) & (revenue_data['Lead Date'] == date) & \
            (revenue_data['s1'] == sub_id)]
    
    revenue = temp_rev_data['Revenue'].sum()
   
    return revenue

def return_cost(revenue_data,date,name,reve_id,sub_id):

    temp_rev_data = revenue_data[
        (revenue_data['Company'] == name) & (revenue_data['Affiliate'] == reve_id) & (revenue_data['Lead Date'] == date) & \
            (revenue_data['s1'] == sub_id)]
    
    cost = temp_rev_data['Cost'].sum()
    
    return cost

def apply_with_progress(df, func, progress_bar):
    total_chunks = 10
    chunk_size = len(df) // total_chunks

    def process_chunk(chunk):
        return chunk.swifter.apply(func, axis=1)

    results = []
    for i in range(total_chunks):
        start = i * chunk_size
        end = (i + 1) * chunk_size if i < total_chunks - 1 else len(df)
        chunk = df.iloc[start:end]
        results.append(process_chunk(chunk))
        progress_bar.progress((i + 1) / total_chunks)

    return pd.concat(results)

def generate_report(input_data_path):

    # Getting the month dictionary
    months_dict = {
        'Jan': 'January',
        'Feb': 'February',
        'Mar': 'March',
        'Apr': 'April',
        'May': 'May',
        'June': 'June',
        'July': 'July',
        'Aug': 'August',
        'Sep': 'September',
        'Oct': 'October',
        'Nov': 'November',
        'Dec': 'December'
    }

    months_dict_2 = {
        "01": "January",
        "02": "February",
        "03": "March",
        "04": "April",
        "05": "May",
        "06": "June",
        "07": "July",
        "08": "August",
        "09": "September",
        "10": "October",
        "11": "November",
        "12": "December"
    }

    # Reading the revenue data
    xls = pd.read_excel(input_data_path, sheet_name=None)
    data_frames = {}

    # xls is a dictionary where the keys are the sheet names and the values are the DataFrames
    for sheet_name, sheet_df in xls.items():
        data_frames[sheet_name] = sheet_df

    # Getting the month name from the comprehensive report data
    month_names = [date.split(' ')[0] for date in data_frames.keys()]
    month_name = (months_dict[list(set(month_names))[0]])

    list_of_campaigns = []

    for key in data_frames.keys():

        cut_off_1 = data_frames[key][data_frames[key]['Campaign'] == 'Publisher Name'].index[0]
        temp_df = data_frames[key].iloc[cut_off_1:]
        cut_off_2 = temp_df[temp_df['Campaign'].isna() == True].index[0]
        campaigns = list(set(list(data_frames[key].iloc[cut_off_1:cut_off_2]['Campaign'].values)))
        campaigns.remove('Publisher Name')

        list_of_campaigns += campaigns

    campaigns = list(set(list_of_campaigns))

    report_data = {

    }

    for campaign in campaigns:

        temp_campaign_df = {

        }

        for key in data_frames.keys():

            temp_df = data_frames[key][data_frames[key]['Campaign'] == campaign]
            useful_columns = ['Publisher', 'Campaign', 'Leads', 'Revenue', 'Unnamed: 4', 'Unnamed: 5','Unnamed: 6']
            temp_df = temp_df[useful_columns]
            temp_df['Date'] = [key] * temp_df.shape[0]
            temp_df.rename(
                columns = {
                    'Unnamed: 4' : 'Clicks/Views', 
                    'Unnamed: 5' : 'We Pay',
                    'Unnamed: 6' : 'Margin'
                },
                inplace=True
            )

            temp_df = temp_df[
                ['Date',
                'Publisher',
                'Campaign',
                'Leads',
                'Revenue',
                'Clicks/Views',
                'We Pay',
                'Margin']
            ]

            # Convert the 'Margin' column to percentage values
            temp_df['Margin'] = temp_df['Margin'].apply(lambda x: f'{int(x*100)}%' if pd.notnull(x) else '0.0%')

            temp_campaign_df[key] = temp_df

        temp_dfs = []

        for key,value in temp_campaign_df.items():

            temp_df = value.copy()

            sub_ids = [id for id in list(set(temp_campaign_df[key]['Publisher'].values)) if '/' in str(id)]

            grouped_dfs = []
            
            try:

                for id in list(set(sub_ids)):

                    grouped_dfs.append(temp_df[temp_df['Publisher'] == id])

            

                temp_df = pd.concat(grouped_dfs,axis = 0)

            except:

                pass

            temp_dfs.append(value)

        blank_row = pd.DataFrame([[np.nan] * temp_df.shape[1]], columns=temp_df.columns)

        # Initialize an empty list to hold the DataFrames with blank rows in between
        concatenated_dfs = []

        # Loop through the DataFrames and concatenate with a blank row in between
        for df in temp_dfs:
            concatenated_dfs.append(df)
            concatenated_dfs.append(blank_row)

        temp_final_df = pd.concat(concatenated_dfs,axis = 0)    

        report_data[campaign] = temp_final_df

    final_report_df = {}

    for key in report_data.keys():

        grouped_dfs = []

        sub_ids = [id for id in list(set(report_data[key]['Publisher'].values)) if '/' in str(id)]

        temp_df = report_data[key]

        for id in list(set(sub_ids)):

            # grouped_dfs.append(temp_df[temp_df['Publisher'] == id])

            total_leads = temp_df[temp_df['Publisher'] == id]['Leads'].sum()
            total_revenue = temp_df[temp_df['Publisher'] == id]['Revenue'].sum()
            total_clicks = temp_df[temp_df['Publisher'] == id]['Clicks/Views'].sum()
            total_cost = temp_df[temp_df['Publisher'] == id]['We Pay'].sum()

            try:
                total_margin = f'{round(((total_revenue - total_cost) / total_revenue)*100,0)}%'
            except:
                total_margin = '0.0%'

            total_data = [month_name,id,key,total_leads,total_revenue,total_clicks,total_cost,total_margin]

            total_data_df = pd.DataFrame([total_data],
                                         columns=temp_df.columns)
            
            total_data_df.rename(
            columns = {
                'Date' : 'Month'
            },
            inplace = True
        )
            grouped_dfs.append(total_data_df) 

            grouped_dfs.append(blank_row)

            

        try: temp_df = pd.concat(grouped_dfs,axis = 0)
        except: pass

        final_report_df[key] = temp_df

    # Step 2: Sort the keys
    sorted_keys = sorted(final_report_df.keys())

    # Step 3: Create a new dictionary with sorted keys
    sorted_dict = {key: final_report_df[key] for key in sorted_keys}

    final_report_df = sorted_dict

    return final_report_df

def get_report(click_reg_path,zip_file_path):

    # Creating an empty directory to save all the findings
    temp_dir = 'synthesized_data/temp_dir/'
    if not os.path.exists(temp_dir): os.mkdir(temp_dir)

    # Checking if there is a zipped file and deleting it
    if os.path.exists('synthesized_data/files.zip'):
        os.remove('synthesized_data/files.zip')

    # Reading the data
    df = read_excel_file(click_reg_path)

    # Getting the all affiliate ids and getting only those affiliates with more than 500 clicks
    useful_affiliates = {}
    for affiliate_id in list(set(df['Affiliate ID'].values)):
        total_clicks = df[df['Affiliate ID'] == affiliate_id]['Cake Clicks (All Clicks)'].sum()
        if total_clicks > 500: 
            useful_affiliates[affiliate_id] = total_clicks

    # Sorting the affiliates based on the number of clicks they have
    useful_affiliates = dict(sorted(useful_affiliates.items(),key = lambda item:item[1],reverse = True))
    affiliate_names = {}
    for affiliate_id in useful_affiliates.keys():
        affiliate_name = list(set(df[df['Affiliate ID'] == affiliate_id]['Affiliate Name'].values))[0]
        affiliate_names[f'{affiliate_name} - {affiliate_id}'] = useful_affiliates[affiliate_id]

    # Finalizing the affiliate names with their respective ids
    useful_affiliates = list(affiliate_names.keys())
    useful_affiliate_ids = [float(id.split('-')[-1].strip()) for id in useful_affiliates]
    useful_affiliate_ids

    # Taking only the useful ids that has more than 500 clicks
    df = df[df['Affiliate ID'].isin(useful_affiliate_ids)]

    df['Percentage'] = round((df['User Registration'] / df['Cake Clicks (All Clicks)']) * 100 , 2)

    # So as we can see S2,S3,S4,S5 contains all null values so we don't need those columns
    useful_columns = [column for column in df.columns if column not in ['S2','S3','S4','S5']]
    df = df[useful_columns]

    # Cleaning the data by removing the null values
    df.dropna(inplace=True)

    # Removing the s2= and s2&#61; value from the s1 columns
    df['S1'] = df['S1'].apply(lambda s1:str(s1).replace('s2','').replace('=','').replace('&#61;',''))

    # Create Streamlit progress bars with text
    st.text('Overall publisher progress...')
    overall_progress_bar = st.progress(0)

    # Getting all the publishers
    publishers = [publisher for publisher in list(set(df['Affiliate Name'].values))]

    df['Percentage'] = df['Percentage'].apply(lambda perc:perc if perc != np.inf else 0)

    total_publishers = len(publishers)
    progress = 0

    # Itterating through all the publishers and getting their data    
    for publisher in publishers:

        # Creating a dictionary containing a dictionary to contain all the dfs
        data_frames = {}

        print(f'------Doing the analysis for {publisher}------')

        # Getting the publishers data frame
        df_for_analysis = df[df['Affiliate Name'] == publisher]

        # Getting the pattern for allinbox
        if publisher == 'All Inbox(Jason Jacobs)':
            df_for_analysis['Revenue Tracker S1'] = df_for_analysis['S1'].apply(lambda s1:s1[5:11])
            df_for_analysis['Prefix S1'] = df_for_analysis['S1'].apply(lambda s1:s1[:5])
            df_for_analysis['Date S1'] = df_for_analysis['S1'].apply(lambda s1:s1[11:])
            df_for_analysis = df_for_analysis[[
                'Date', 'Affiliate Name', 'Affiliate ID', 'Revenue Tracker ID', 'S1','Prefix S1','Revenue Tracker S1','Date S1',
                'User Registration', 'Cake Clicks (All Clicks)', 'Percentage'
            ]]

        data_frames['Original File'] = df_for_analysis

        # Getting the rows with error ids
        error_df = df_for_analysis[(df_for_analysis['Cake Clicks (All Clicks)'] == 0) & (df_for_analysis['User Registration'] > 0)]
        error_df = error_df.sort_values(by = 'S1',ascending = False)

        # Getting abnormal sub ids
        abnormal_df = df_for_analysis[(df_for_analysis['User Registration'] > df_for_analysis['Cake Clicks (All Clicks)']) & \
                                    (df_for_analysis['Cake Clicks (All Clicks)'] != 0)]

        # Now let us take the clean df that contain correct sub ids
        df_for_analysis = df_for_analysis[(df_for_analysis['Cake Clicks (All Clicks)'] != 0) & \
                                        (df_for_analysis['User Registration'] <= df_for_analysis['Cake Clicks (All Clicks)'])]
        
        # st.dataframe(df_for_analysis)

        # Now let us see how many duplicate sub ids we have
        temp_df = pd.DataFrame(
            df_for_analysis['S1'].value_counts()
        )
        temp_df.reset_index(inplace=True)
        temp_df.rename(
            columns = {
                'index' : 'S1',
                'count' : 'Number of Repetition'
            },
            inplace = True
        )

        duplicated_sub_ids = temp_df[temp_df['Number of Repetition'] > 1]['S1'].values

        if publisher == 'All Inbox(Jason Jacobs)':

            data = {
                'Date S1' : [],
                'Day S1' : [],
                'Month S1' : [],
                'Year S1' : [],
                'User Registration' : [],
                'Cake Clicks (All Clicks)' : [],
                'Average Percentage' : []
            }

            for id in list(set(df_for_analysis['Date S1'].values)):
                
                total_reg = df_for_analysis[df_for_analysis['Date S1'] == id]['User Registration'].sum()
                total_clicks = df_for_analysis[df_for_analysis['Date S1'] == id]['Cake Clicks (All Clicks)'].sum()
                # percentage = round((total_reg/total_clicks) * 100 , 2)
                percentage = np.mean(list(df_for_analysis[df_for_analysis['Date S1'] == id]['Percentage'].values))

                data['Date S1'].append(id)
                data['Day S1'].append(id[4:])
                data['Month S1'].append(id[2:4])
                data['Year S1'].append(id[:2])
                data['User Registration'].append(total_reg)
                data['Cake Clicks (All Clicks)'].append(total_clicks)
                data['Average Percentage'].append(percentage)

            pattern_df = pd.DataFrame(data)
            pattern_df = pattern_df.sort_values(by = ['Average Percentage','User Registration'],ascending = False)

            data_frames['Pattern (Date)'] = pattern_df

            data_prefix = {
                'Prefix S1' : [],
                'User Registration' : [],
                'Cake Clicks (All Clicks)' : [],
                'Average Percentage' : []
            }

            for id in list(set(df_for_analysis['Prefix S1'])):
                
                total_reg = df_for_analysis[df_for_analysis['Prefix S1'] == id]['User Registration'].sum()
                total_clicks = df_for_analysis[df_for_analysis['Prefix S1'] == id]['Cake Clicks (All Clicks)'].sum()
                # percentage = round((total_reg/total_clicks) * 100 , 2)
                percentage = np.mean(list(df_for_analysis[df_for_analysis['Prefix S1'] == id]['Percentage'].values))

                data_prefix['Prefix S1'].append(id)
                data_prefix['User Registration'].append(total_reg)
                data_prefix['Cake Clicks (All Clicks)'].append(total_clicks)
                data_prefix['Average Percentage'].append(percentage)

            prefix_pattern_df = pd.DataFrame(data_prefix) 
            prefix_pattern_df = prefix_pattern_df.sort_values(by = ['Average Percentage','User Registration'],ascending = False)

            data_frames['Pattern (Prefix)'] = prefix_pattern_df

        # Let us create a new column letting us know if the id is duplicated or not
        df_for_analysis['Duplicate'] = df_for_analysis['S1'].apply(lambda s1 : 'Yes' if s1 in duplicated_sub_ids else 'No')

        # Finding the duplicated and not duplicated dataframe
        dup_df = df_for_analysis[df_for_analysis['Duplicate'] == 'Yes']
        data_frames['Dup_Sub_IDs'] = dup_df
        no_dup_df = df_for_analysis[df_for_analysis['Duplicate'] == 'No']
        no_dup_df['Average Percentage'] = no_dup_df['Percentage']
        data_frames['Not_Dup_Sub_IDs'] = no_dup_df
        non_dup_sorted_by_percentage = no_dup_df.sort_values(by = ['Average Percentage','S1'],ascending = False)

        # Getting the duplicated sub ids that has more than 40% registration rate
        dup_more_40 = dup_df[(dup_df['Percentage'] > 40) & (dup_df['Percentage'] != np.inf)]
        data_frames['Dup_Sub_IDs>40%_regi_rate'] = dup_more_40

        # Getting the duplicated sub ids that has less than 40% registration rate
        dup_less_40 = dup_df[
            (dup_df['Percentage'] >= 11) & 
            (dup_df['Percentage'] <= 40) & 
            (dup_df['Percentage'] != np.inf)
            ]
        data_frames['Dup_Sub_IDs<40%_nd>10%'] = dup_less_40

        # Getting the duplicated sub ids that has exactly 40% registration rate
        dup_exactly_40 = dup_df[(dup_df['Percentage'] <= 10) & (dup_df['Percentage'] != np.inf)]
        data_frames['Dup_Sub_IDs<10%_regi_rate'] = dup_exactly_40

        # Getting the non duplicated sub ids and arranging them
        data_frames['Non-Dup_Sub_IDs_sorted_by_perc'] = non_dup_sorted_by_percentage

        # Getting the non duplicated sub ids that has more than 40% registration rate
        non_dup_more_40 = no_dup_df[(no_dup_df['Percentage'] > 40) & (no_dup_df['Percentage'] != np.inf)]
        data_frames['Not-Dup_Sub_IDs>40%_reg_rate'] = non_dup_more_40

        # Getting the non duplicated sub ids that has less than 40% registration rate
        # non_dup_less_40 = no_dup_df[(no_dup_df['Percentage'].isin(list(range(11,41)))) & (no_dup_df['Percentage'] != np.inf)]
        non_dup_less_40 = no_dup_df[(no_dup_df['Percentage'] >= 10) & 
                            (no_dup_df['Percentage'] <= 40) & 
                            (no_dup_df['Percentage'] != np.inf)]
        data_frames['Not-Dup_Sub_IDs<40%_nd>10%'] = non_dup_less_40

        # Getting the non duplicated sub ids that has exactly 40% registration rate
        non_dup_exactly_40 = no_dup_df[(no_dup_df['Percentage'] < 10) & (no_dup_df['Percentage'] != np.inf)]
        data_frames['Not-Dup_Sub_IDs<10%_reg_rate'] = non_dup_exactly_40

        # Making each duplicated sub ids together
        dup_sorted_df = dup_df.sort_values(by=['Duplicate', 'S1'],ascending = False)

        # Getting the average value for each duplicated sub ids
        def get_average_registration_rate(id):

            total_clicks = dup_sorted_df[dup_sorted_df['S1'] == id]['Cake Clicks (All Clicks)'].sum()

            total_registration = dup_sorted_df[dup_sorted_df['S1'] == id]['User Registration'].sum()

            # average_reg_rate = round((total_registration/total_clicks) * 100 , 2)

            average_reg_rate = np.mean(list(df_for_analysis[df_for_analysis['S1'] == id]['Percentage'].values))
    
            return total_clicks,total_registration,average_reg_rate
        
        registration_mapping = {}

        reg_mapping = {}

        click_mapping = {}


        for id in duplicated_sub_ids:

            total_clicks,total_registration,avg_reg_rate = get_average_registration_rate(id)
            registration_mapping[id] = avg_reg_rate
            reg_mapping[id] = total_registration
            click_mapping[id] = total_clicks


        # Sorting the sub ids
        dup_sorted_df['Average Percentage'] = dup_sorted_df['S1'].apply(lambda s1:registration_mapping[s1])

        # Doing duplicate sorting analysis based on average registration rate 
        dup_sorted_df_by_reg_rate = dup_sorted_df.sort_values(by = ['Average Percentage','S1'],ascending = False)
        temp_dfs = []
        temp_list_id = []
        for id in list(dup_sorted_df_by_reg_rate['S1'].values):

            if id not in temp_list_id:

                temp_dfs.append(dup_sorted_df_by_reg_rate[dup_sorted_df_by_reg_rate['S1'] == id])

                if publisher == 'All Inbox(Jason Jacobs)':
                    total_data = ['','','','',
                                'Total','','','',
                                reg_mapping[id],
                                click_mapping[id],
                                '',
                                '',
                                registration_mapping[id]
                                ]
                    
                else:
                    total_data = ['','','','',
                                'Total',
                                reg_mapping[id],
                                click_mapping[id],
                                '',
                                '',
                                registration_mapping[id]
                                ]

                # st.warning(f'{total_data}-{publisher}-{list(dup_sorted_df_by_reg_rate.columns)}')
                total_data_df = pd.DataFrame([total_data],
                                                columns=dup_sorted_df_by_reg_rate.columns)
                
                temp_dfs.append(total_data_df)

            temp_list_id.append(id)

        if len(temp_dfs) > 0:
            dup_sorted_df_by_reg_rate = pd.concat(temp_dfs,axis = 0)
        else:
            pass
             
        data_frames['Duplicated_Sorted_by_avg_reg_rate'] = dup_sorted_df_by_reg_rate

        dup_must_stop = dup_sorted_df_by_reg_rate[dup_sorted_df_by_reg_rate['Average Percentage'] == 0]
        non_dup_must_stop = non_dup_sorted_by_percentage[non_dup_sorted_by_percentage['Average Percentage'] == 0]
        average_perc_0_reg_rate = pd.concat(
            [dup_must_stop,non_dup_must_stop],
            axis = 0
        )

        if publisher == 'All Inbox(Jason Jacobs)':

            top_10 = {

            }

            for tracker_id in [7750,8866,8867,8895]:

                all_in_box = df_for_analysis[df_for_analysis['Revenue Tracker ID'] == tracker_id]
                
                if tracker_id in [7750,8867]:

                    all_in_box = all_in_box.sort_values('User Registration',ascending = False)

                    grouped_df = all_in_box.groupby(by = 'S1').agg('sum')

                    grouped_df.sort_values('User Registration',ascending = False,inplace = True)

                    check_df = grouped_df.reset_index()
                    check_df = check_df[
                        [
                            'S1',
                            'User Registration',
                            'Cake Clicks (All Clicks)'

                        ]
                    ]

                    check_df['Average Percentage'] = check_df['User Registration']  / check_df['Cake Clicks (All Clicks)']

                    check_df['Average Percentage'] = check_df['Average Percentage'].apply(lambda perc:round(perc*100,2))
                    
                    top_10[f'Top_10_Sub_IDs_{tracker_id}'] = check_df.head(10)

                else:

                    grouped_df = all_in_box.groupby('S1').agg('sum')

                    average_cake_clicks = grouped_df['Cake Clicks (All Clicks)'].mean()

                    grouped_df = grouped_df[grouped_df['Cake Clicks (All Clicks)'] > average_cake_clicks]

                    grouped_df.reset_index( inplace = True )

                    check_df = grouped_df [
                        [
                            'S1',
                            'User Registration',
                            'Cake Clicks (All Clicks)'
                        ]
                    ]

                    check_df['Average Percentage'] = check_df['User Registration']  / check_df['Cake Clicks (All Clicks)']

                    check_df['Average Percentage'] = check_df['Average Percentage'].apply(lambda perc:round(perc*100,2))

                    check_df.sort_values('Average Percentage',ascending = False, inplace = True)
                    
                    top_10[f'Top_10_Sub_IDs_{tracker_id}'] = check_df.head(10)

            
            for key in top_10.keys():

                data_frames[key] = top_10[key]

        data_frames['Abnormal Sub IDs'] = abnormal_df

        data_frames['Avg_perc_0%'] = average_perc_0_reg_rate

        data_frames['Errors'] = error_df

        dup_must_stop = dup_sorted_df_by_reg_rate[(dup_sorted_df_by_reg_rate['Percentage'] == 0) &
                                                  (dup_sorted_df_by_reg_rate['Cake Clicks (All Clicks)'] > 20)]
        non_dup_must_stop = non_dup_sorted_by_percentage[(non_dup_sorted_by_percentage['Percentage'] == 0)&
                                                  (non_dup_sorted_by_percentage['Cake Clicks (All Clicks)'] > 20)]
        must_stop_df = pd.concat(
            [dup_must_stop,non_dup_must_stop],
            axis = 0
        )

        data_frames['Must Stop 0%'] = must_stop_df

        data_frames['Conclusion'] = pd.DataFrame()

        with pd.ExcelWriter(f'{temp_dir}{publisher}.xlsx', engine='openpyxl') as writer:
            for sheet_name, saving_df in data_frames.items():

                saving_df.to_excel(writer, index = False, sheet_name=sheet_name)

                # Access the workbook and worksheet
                workbook = writer.book
                worksheet = workbook[sheet_name]

                conditional_formating = False
                error_sheet = False
                warning_sheet = False
                findings = False

                # Apply styles
                if publisher != 'All Inbox(Jason Jacobs)':
                    top_10 = {}

                if sheet_name in ['Non-Dup_Sub_IDs_sorted_by_perc','Duplicated_Sorted_by_avg_reg_rate','Pattern (Date)','Pattern (Prefix)'] or \
                    sheet_name in list(top_10.keys()):
                    
                    conditional_formating = 'registration_rate'

                if sheet_name in ['Non-Dup_Sub_IDs_sorted_by_margin','Duplicated_Sorted_by_avg_margin','Pattern (Surfix)']: 
                    conditional_formating = 'margin'
                    
                if sheet_name in ['Must Stop 0%','Errors','Avg_perc_0%'] : error_sheet = True

                if sheet_name == 'Abnormal Sub IDs' : warning_sheet = True

                if sheet_name == 'Conclusion' : 

                    below_40_nd_above_10 = len(
                        list(set(dup_sorted_df_by_reg_rate[(dup_sorted_df_by_reg_rate['Average Percentage'] >= 10) & \
                                                           (dup_sorted_df_by_reg_rate['Average Percentage'] <= 40)]['S1'].values))) + \
                            len(list(set(non_dup_less_40['S1'].values))) - 1

                    abv_40 = len(
                        list(set(dup_sorted_df_by_reg_rate[dup_sorted_df_by_reg_rate['Average Percentage'] > 40]['S1'].values))) + \
                            len(list(set(non_dup_more_40['S1'].values))) - 1

                    below_10 = len(
                        list(set(dup_sorted_df_by_reg_rate[dup_sorted_df_by_reg_rate['Average Percentage'] < 10]['S1'].values))) + \
                            len(list(set(non_dup_exactly_40['S1'].values))) - 1
                    
                    unique_ids = len(list(set(df_for_analysis['S1'].values)))
                   
                    must_stop = len(list(set(must_stop_df['S1'].values)))
                    
                    errors = len(list(set(error_df['S1'].values)))

                    findings = [
                        f'We have {unique_ids} unique s1',
                        f'We have {0 if must_stop in [1,-1] else must_stop} s1 Must stop ids 0% ',
                        f'{below_10} below 10%',
                        f'{below_40_nd_above_10} Below 40 % and above 10%',
                        f'{abv_40} good s1 above 40%',
                        f'{errors} error s1'
                    ]

                format_sheet(worksheet,conditional_formating,error_sheet,warning_sheet,findings)

        progress += 1
        overall_progress_bar.progress(progress / total_publishers)

        print(f'***Finished saving the file***')

    print('////////////////Zipping the files////////////////')
    # Zipping the already created dataframes
    files = [temp_dir + file for file in os.listdir(temp_dir)]
    zip_files(files,zip_file_path)

    # Removing the temp directory
    shutil.rmtree(temp_dir)
